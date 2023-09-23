
from openpyxl import load_workbook


# The output to be written is a list of ints, saying what technician do what job.
def write_output(output : list[int], excel_path : str):
    excelFile = load_workbook(filename = excel_path, data_only=True)
    print("\n\n\n\n-----")
    print(excelFile.sheetnames)

    output_sheet = excelFile.create_sheet("Output")
    output_sheet['A1'] = 'Olá'
    excelFile.save(excel_path)
