from openpyxl import load_workbook

# Tutorial
# https://www.pylenin.com/blogs/excel-with-python/
wbFile = load_workbook(filename = 'input.xlsx', data_only=True)

cell_n_projects = "C1"
cell_n_tecns = "C2"
start_tecn_row = 6
start_tecn_col = 3

if "Compatibilidade" in wbFile.sheetnames:
    sheet = wbFile["Compatibilidade"]
    n_tecns = sheet[cell_n_tecns].value
    n_projects = sheet[cell_n_projects].value
    print("Numbers")
    print(n_projects)
    print(n_tecns)
    for current_tecn in range(n_tecns):
        current_row = start_tecn_row+current_tecn
        for current_project in range(n_projects):
            current_col = start_tecn_col+current_project
            print(f"Aptidão tecn: {current_tecn+1} | proj : {current_project+1} : {sheet.cell(row=current_row, column=current_col).value}")

