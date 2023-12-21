
from openpyxl import load_workbook

# The id and other stuff are hardcoded, but they shouldn't...
# Because this is temporary, and the effort to change is high.
def read_translation(sheetname : str, workbook, col_id : int, col_value : int):
    res_list : list(str) = [None]
    tecs_names = workbook.get_sheet_by_name(sheetname)
    for row in tecs_names.iter_rows():
        tecn_id = row[col_id].value
        tecn_name = row[col_value].value
        try:
            res_list.insert(int(tecn_id), tecn_name)
            print(int(tecn_id), " - ", tecn_name)
        except:
            pass
    return res_list

def read_names(path_excel_input : str):
    excelFile = load_workbook(filename = path_excel_input, data_only=True)
    translation_tecns = read_translation(sheetname="Tecnicos", workbook=excelFile, col_id=0, col_value=2 )
    translation_projs = read_translation(sheetname="Projetos", workbook=excelFile, col_id=0, col_value=17)
    return (translation_tecns, translation_projs)
