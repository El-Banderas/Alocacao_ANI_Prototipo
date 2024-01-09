from configExcel import *
import pandas as pd
import openpyxl 
from CarolUpdateRemainingCost import carolina_heuristicaURC
from Gui_heuristica import gui_heuristica
import itertools
import statistics 
from datetime import timedelta

######################################### READ EXCEL ################################################
# load excel with its path 
filePath = "ZéPaulo.xlsx"
wrkbk = openpyxl.load_workbook(filePath, data_only=True) 

sh_projs = wrkbk["Output AS-IS"] 

projetos_organizados_por_meses = {}

# ID -> Cost anal
costs_anal = {}
# ID -> Cost accomp
costs_accomp = {}

# Key -> Proj id , value -> tecn_anal
projetos_alocacoes = {}

VERY_HIGH_VALUE = 10000

num_techs = 5

previous_costs_carol = [0] * num_techs
#previous_costs_gui = [0] * num_techs

# iterate through excel and display data 
for row in sh_projs.iter_rows(min_row=2): 
    month_anal = row[pos_init_date].value + timedelta(days=2)#.strftime("%m/%d/%Y")
    month_accomp = row[pos_accomp_date].value + timedelta(days=2)#.strftime("%m/%d/%Y")
    # Matrix row in costs starts in 0, index in this sheet begins in 1
    row_proj = row[pos_id].value 
    costs_anal[row_proj] = row[pos_cost_an].value
    costs_accomp[row_proj] = row[pos_cost_accomp].value
    if month_anal in projetos_organizados_por_meses:
        projetos_organizados_por_meses[month_anal].append(f"A{row_proj}")
    else:
        projetos_organizados_por_meses[month_anal] = [f"A{row_proj}"]
    if month_accomp in projetos_organizados_por_meses:
        projetos_organizados_por_meses[month_accomp].append(f"B{row_proj}")
    else:
        projetos_organizados_por_meses[month_accomp] = [f"B{row_proj}"]
    
matrix_aptitude = []



print("Meses")
#print(projetos_organizados_por_meses)
print(dict(sorted(projetos_organizados_por_meses.items())))
print(costs_anal)
print(costs_accomp)
print("\n\n\n")
######################################### HEURISTIC STUFF ################################################


def print_matrix(mtx):
    print("#####################")
    for row in mtx:
        print(row)
    print("#####################")
# Join current costs to the previous ones
def add_current_to_previous_costs(alocation : list[int],this_matrix_costs : list[list[int]], previous_costs : list[int] ):
    for proj, tech in enumerate(alocation):
        previous_costs[tech] += this_matrix_costs[proj][tech]
        print(f"Add {tech}: {this_matrix_costs[proj][tech]}")
    print(previous_costs)
    return previous_costs

def invert_mtx(mtx : list[list[int]]):
    inv_mtx = []
    for line in mtx:
        new_line = []
        for cell in line:
            new_value = 1 / cell
            new_line.append(new_value)
        inv_mtx.append(new_line)
    return inv_mtx

n_months = 0
total_months = len(projetos_organizados_por_meses.keys())
# Check all projects by month
for month, projects in dict(sorted(projetos_organizados_por_meses.items())).items():
    n_months = n_months + 1
    
    print("Alocation month: ", n_months , " / " , total_months)
    this_matrix = []
    # Convert line matrix to project ID, so heuristic could use 
    # Like, line 3 of matrix is project 12
    conversion_this_month = {}
    # Construct matrix of aptitudes between projects of this month and techs
    for idx, proj in enumerate(projects):
        conversion_this_month[idx] = proj
        # IF analysis project
        if "A" in proj:
            id_proj = int(proj[1:])
            this_cost = [costs_anal[id_proj]] * num_techs
        else: 
            id_proj = int(proj[1:])
            this_cost = [costs_anal[id_proj]] * num_techs
            id_tecn_anal =  projetos_alocacoes[f"A{id_proj}"]
            this_cost[id_tecn_anal] = VERY_HIGH_VALUE
            print(" CHANGE HIGH VALUE (proj id / tecn) : " , id_proj , " / ", id_tecn_anal)
        this_matrix.append(this_cost)
    num_projs = len(this_matrix)
    this_matrix_pd = pd.DataFrame(this_matrix)
    # Do the alocations for two heuristics
    alocation = carolina_heuristicaURC(df_aptitude_between_task_machine=this_matrix_pd, num_tasks=num_projs, machine_quantity=num_techs, previous_costs=previous_costs_carol)
    for idx, tecn_alocated in enumerate(alocation):
        id_proj = conversion_this_month[idx]
        projetos_alocacoes[id_proj] = tecn_alocated
        if "A" in id_proj:
            this_cost = costs_anal[int(id_proj[1:])]
        else:
            this_cost = costs_accomp[int(id_proj[1:])]
        previous_costs_carol[tecn_alocated] += this_cost
    # In case the matrix should be inverted (if heuristic interprets the matrix other way). But it's commented
    # this_matrix_pd = pd.DataFrame(invert_mtx(this_matrix))
    #alocation = gui_heuristica(df_task_time_per_machine=this_matrix_pd, load_per_machine=previous_costs_gui, num_tasks=num_projs, machine_quantity=num_techs)
    # To join list of lists


'''

def print_stuff(total_costs : list[int]):
    print(total_costs)
    print("STD: ",statistics.pstdev(total_costs) )
    print("Amp: ",max(total_costs) - min(total_costs) )

print("Carol")
print_stuff(total_costs=previous_costs_carol)
print("Gui")
print_stuff(total_costs=previous_costs_gui)

print("Num months: ", n_months)
'''
print("THE END")
print(projetos_alocacoes)
print(previous_costs_carol)
print("Writing to excel")


output = wrkbk["Input TO-BE"] 
# Associates the id of project to the line it is written in excel
id_proj_line = {}
current_line = 2
for proj_name, tecn in dict(sorted(projetos_alocacoes.items())).items():
    proj_id = int(proj_name[1:])
    if proj_id in id_proj_line:
        proj_line = id_proj_line[proj_id]
    else:
        proj_line = current_line
        id_proj_line[proj_id] = current_line
        current_line += 1
    if "A" in proj_name:
        cellID= output.cell(row=proj_line, column=1)
        cellID.value = proj_id
        
        cellAnal= output.cell(row=proj_line, column=2)
        cellAnal.value = tecn
    else:
        cellAnal= output.cell(row=proj_line, column=3)
        cellAnal.value = tecn
wrkbk.save(filePath)
wrkbk.close()