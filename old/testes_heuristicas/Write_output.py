from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook


def write_results(excel_workbook,  pairs_with_data, iterations : int):
    output_sheet = excel_workbook.create_sheet("results")
    
    output_sheet['A1'] = 'Stand. dev.'
    output_sheet['D1'] = 'Amp.'
    output_sheet['G1'] = 'Custo total.'
    output_sheet['A2'] = 'Gui'
    output_sheet['B2'] = 'Carolina'
    output_sheet['C2'] = 'Renan'
    output_sheet['D2'] = 'Gui'
    output_sheet['E2'] = 'Carolina'
    output_sheet['F2'] = 'Renan'
    output_sheet['G2'] = 'Gui'
    output_sheet['H2'] = 'Carolina'
    output_sheet['I2'] = 'Renan'
    (results_std, results_total_cost, amplitude, pair_input, attributions) = pairs_with_data
    for iteration in range(iterations):
        #Write Gui
        output_sheet.cell(row=iteration+3, column=1).value = results_std['Gui'][iteration]
        output_sheet.cell(row=iteration+3, column=2).value = results_std['Carolina'][iteration]
        output_sheet.cell(row=iteration+3, column=3).value = results_std['Renan'][iteration]

        output_sheet.cell(row=iteration+3, column=7).value = results_total_cost['Gui'][iteration]
        output_sheet.cell(row=iteration+3, column=8).value = results_total_cost['Carolina'][iteration]
        output_sheet.cell(row=iteration+3, column=9).value = results_total_cost['Renan'][iteration]

        output_sheet.cell(row=iteration+3, column=4).value = amplitude['Gui'][iteration]
        output_sheet.cell(row=iteration+3, column=5).value = amplitude['Carolina'][iteration]
        output_sheet.cell(row=iteration+3, column=6).value = amplitude['Renan'][iteration]

def write_little_table(sheet):
        sheet.cell(row=1, column=2).value = "Stand. dev"
        sheet.cell(row=1, column=3).value = "Custo total"
        sheet.cell(row=1, column=4).value = "Amplitude"
        sheet.cell(row=1, column=5).value = "Atribuições"
        
        sheet.cell(row=2, column=1).value = "Gui"
        sheet.cell(row=3, column=1).value = "Renan"
        sheet.cell(row=4, column=1).value = "Carolina"

        sheet.cell(row=7, column=1).value = "Tamanho previsto trabalhos"

        sheet.cell(row=10, column=1).value = "Matriz aptidões"


def write_input_and_output(excel_workbook, pairs_with_data, iterations : int): 
    (results_std, results_total_cost, amplitude, pair_input, results_attributions) = pairs_with_data
    for iteration in range(iterations):
        output_sheet = excel_workbook.create_sheet(f"Iteration_{iteration}")
        write_little_table(output_sheet)
        output_sheet.cell(row=2, column=2).value = results_std['Gui'][iteration]
        output_sheet.cell(row=2, column=3).value = results_total_cost['Gui'][iteration]
        output_sheet.cell(row=2, column=4).value = amplitude['Gui'][iteration]
        output_sheet.cell(row=2, column=5).value = str(results_attributions['Gui'][iteration])

        output_sheet.cell(row=4, column=2).value = results_std['Carolina'][iteration]
        output_sheet.cell(row=4, column=3).value = results_total_cost['Carolina'][iteration]
        output_sheet.cell(row=4, column=4).value = amplitude['Carolina'][iteration]
        output_sheet.cell(row=4, column=5).value = str(results_attributions['Carolina'][iteration])

        output_sheet.cell(row=3, column=2).value = results_std['Renan'][iteration]
        output_sheet.cell(row=3, column=3).value = results_total_cost['Renan'][iteration]
        output_sheet.cell(row=3, column=4).value = amplitude['Renan'][iteration]
        output_sheet.cell(row=3, column=5).value = str(results_attributions['Renan'][iteration])

        # Trabalhos totais
        (tasks_times, aptitudes) = pair_input[iteration]
        for i in range(len(tasks_times)):
            output_sheet.cell(row=7, column=i+2).value = str(tasks_times.loc[i].values)

        (rows_num, col_num ) = aptitudes.shape
        for row in range(rows_num):
            for col in range(col_num):
                output_sheet.cell(row=11+row, column=col+1).value = aptitudes.iat[row, col]
             

def write_one_excel(name_excel : str, info_pair, interations : int):
    output_excel = Workbook()
    write_results(excel_workbook=output_excel, pairs_with_data=info_pair, iterations=interations) 
    write_input_and_output(excel_workbook=output_excel, pairs_with_data=info_pair, iterations=interations) 
    output_excel.save(f"./{name_excel}.xlsx")


# The output to be written is a list of ints, saying what technician do what job.
def write_output(pair_small, pair_medium, pair_large, iteration : int):
 
    print("Starting writing output")

    write_one_excel(name_excel="SMALL", info_pair=pair_small, interations=iteration) 
    write_one_excel(name_excel="MEDIUM", info_pair=pair_medium, interations=iteration) 
    write_one_excel(name_excel="LARGE", info_pair=pair_large, interations=iteration) 

