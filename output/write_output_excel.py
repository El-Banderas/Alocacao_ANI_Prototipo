from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def write_attribution(excel_workbook, attribution : list[int]):
    output_sheet = excel_workbook.create_sheet('Atribuicoes')
    
    output_sheet['A1'] = 'Atribuição de técnicos a projetos'
    output_sheet['A2'] = 'Tarefas'
    output_sheet['A3'] = 'Técnicos atribuídos'
    for task in range(len(attribution)):
        output_sheet.cell(row=2, column=task+2).value = task+1
        output_sheet.cell(row=3, column=task+2).value = attribution[task]+1
        output_sheet.cell(row=2, column=task+2).border = thin_border
        output_sheet.cell(row=3, column=task+2).border = thin_border


# The output to be written is a list of ints, saying what technician do what job.
def write_output(output : list[int], excel_path : str):
 
    # output_excel  = load_workbook(filename = excel_path)
    output_excel = Workbook()
    
    write_attribution(excel_workbook=output_excel, attribution=output)

    # Delete default sheet
    
    output_excel.save(excel_path)
