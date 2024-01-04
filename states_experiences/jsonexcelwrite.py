import openpyxl


def write_projs_to_excel(self, filename="und.xlsx", worksheet_name="Tecnicos"):
    wb = openpyxl.load_workbook(filename)
    ws = wb[worksheet_name]

    # headers = ["Tech ID", "Tech Name", "Tech Specialization"]
    # for col_num, header in enumerate(headers, 1):
    #     ws.cell(row=1, column=col_num, value=header)

    # Write tech information
    for row_num, tech in enumerate(self.tecns, 2):
        ws.cell(row=row_num, column=1, value=tech.id)
        ws.cell(row=row_num, column=3, value=tech.name)
        ws.cell(
            row=row_num, column=11, value=tech.specialization
        )  # mudar para Data de afiliação , Disponibilidade?

    # Save the workbook to the specified file
    wb.save(filename)


def write_projs_to_excel(self, filename="und.xlsx", worksheet_name="Projetos"):
    wb = openpyxl.load_workbook(filename)
    ws = wb[worksheet_name]

    # headers = ["Project ID", "Área", "Fase atual", "Técnico de análise", "Gestor de projeto", "Tipo de projeto", "Data Inicio", "Data Fim", "Sigla"]
    # for col_num, header in enumerate(headers, 1):
    #    ws.cell(row=1, column=col_num, value=header)

    for row_num, proj in enumerate(self.projs, 2):
        ws.cell(row=row_num, column=1, value=proj.id)
        ws.cell(row=row_num, column=2, value=proj.area)
        ws.cell(row=row_num, column=3, value=proj.current_phase)
        ws.cell(row=row_num, column=4, value=proj.analysis_tech)
        ws.cell(row=row_num, column=5, value=proj.project_manager)
        ws.cell(row=row_num, column=8, value=proj.project_type)
        ws.cell(row=row_num, column=16, value=proj.start_date)
        ws.cell(row=row_num, column=17, value=proj.end_date)
        ws.cell(row=row_num, column=18, value=proj.sigla)

    wb.save(filename)
