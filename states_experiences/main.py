
from openpyxl import load_workbook
from Proj import Project
from Tec import Tech, read_tech
import random

path = "./states.xlsx"

projs_to_alocate_analisis = set()
projs_to_alocate_manager = set()
projs_closed = set()
projs_already_alocated = set()


def read_projects(workbook):
    projetos_sheet = workbook["Projetos"]
    current_line = 2

    while (True):
        this_id = projetos_sheet.cell(row=current_line, column=1).value
        if this_id == None:
            break
        this_area = projetos_sheet.cell(row=current_line, column=2).value
        this_current_phase = projetos_sheet.cell(row=current_line, column=3).value
        this_tec_anal = projetos_sheet.cell(row=current_line, column=4).value
        this_tec_manager = projetos_sheet.cell(row=current_line, column=5).value
        this_proj = Project(id=this_id, area=this_area, current_phase=this_current_phase, tec_anal=this_tec_anal, tec_manager=this_tec_manager, line_index=current_line)
        # project alocated
        if this_proj.current_phase == 2:
            projs_closed.add(this_proj)
        # Project not approved 
        elif this_proj.current_phase == 0 :
            # Project not approved, and no tec is reviewing it
            if this_proj.tec_anal == None:
                projs_to_alocate_analisis.add(this_proj)
            # Project is being reviewed
            else:
                projs_already_alocated.add(this_proj)
        elif this_proj.current_phase == 1:
            if this_proj.tec_manager == None:
                assert(this_proj.tec_anal != None)
                pair_black_list = (this_proj, this_proj.tec_anal)
                projs_to_alocate_manager.add(pair_black_list)
            else:
                projs_already_alocated.add(this_proj)
        else:
            print("ERROR INVALID STATE: ", this_proj.current_phase)
            raise Exception(f"Invalid state of project {this_proj.id}.")
        current_line+=1 

def print_list(list):
    for x in list:
        print(x, end=' ; ')
    print()

def make_atribution(sheet_tecns, sheet_projects, list_techs):
    for project in projs_to_alocate_analisis:
        random_choosen_tec = random.randint(1, len(list_techs))
        # Write tech to the project atribution
        cellref=sheet_projects.cell(row=project.line_index, column=4)
        cellref.value= random_choosen_tec
        
        selected_tech = list(filter(lambda x: (x.id == random_choosen_tec), list_techs))[0] 
        cellref=sheet_tecns.cell(row=selected_tech.line_index, column=3)
        cellref.value= f"{cellref.value} ; Análise{project.id}"


    for (project, no_tech) in projs_to_alocate_manager:
        random_choosen_tec = random.randint(1, len(list_techs))
        while random_choosen_tec == no_tech:
            random_choosen_tec = random.randint(1, len(list_techs))

        cellref=sheet_projects.cell(row=project.line_index, column=5)
        cellref.value= random_choosen_tec

        selected_tech = list(filter(lambda x: (x.id == random_choosen_tec), list_techs))[0] 
        cellref=sheet_tecns.cell(row=selected_tech.line_index, column=3)
        cellref.value= f"{cellref.value} ; Acompanhamento{project.id}"


if __name__ == '__main__':
    wb_obj = load_workbook(path)
    techs_list = read_tech(workbook=wb_obj)
    read_projects(workbook=wb_obj)
    tecns = wb_obj["Técnicos"]

    if len(projs_already_alocated) > 0:
        print("Projectos alocados")
        print_list(projs_already_alocated)

    if len(projs_to_alocate_analisis) > 0:
        print("Projects precisam tec candidatura")
        print_list(projs_to_alocate_analisis)
    if len(projs_to_alocate_manager) > 0:
        print("Projects precisam tec acompanhamento")
        print_list(projs_to_alocate_manager)
    if len(projs_closed) > 0:
        print("Projects feitos")
        print_list(projs_closed)
    make_atribution( sheet_tecns=wb_obj["Técnicos"], sheet_projects=wb_obj["Projetos"], list_techs=techs_list)
    wb_obj.save(path)
    