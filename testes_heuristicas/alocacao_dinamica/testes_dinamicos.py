from configExcel import *
import pandas as pd
import openpyxl 
from CarolUpdateRemainingCost import carolina_heuristicaURC
from Gui_heuristica import gui_heuristica
import itertools
import statistics 

# load excel with its path 
wrkbk = openpyxl.load_workbook("Dados Exemplos.xlsx", data_only=True) 

sh_projs = wrkbk["Candidaturas"] 

projetos_organizados_por_meses = {}

# iterate through excel and display data 
for row in sh_projs.iter_rows(min_row=3, min_col=1): 
    month_this_line = row[pos_mes].value
    # Matrix row in costs starts in 0, index in this sheet begins in 1
    row_proj = row[pos_id].value -1
    cost = row[pos_cost].value
    if month_this_line in projetos_organizados_por_meses:
        projetos_organizados_por_meses[month_this_line].append(row_proj)
    else:
        projetos_organizados_por_meses[month_this_line] = [row_proj]

sh_matrix = wrkbk["Matriz Total"] 
print("\n\n\nMatriz total")
matrix_aptitude = []
for row in sh_matrix.iter_rows(min_row=3, min_col=2): 
    if row[0].value != None:
        this_row = []
        for tecn in range(1,6):
            this_value = row[tecn].value
            this_row.append(this_value)        
        matrix_aptitude.append(this_row)

months_init_projs = list(projetos_organizados_por_meses.keys())
months_init_projs.sort()

num_techs = len(matrix_aptitude[0])
previous_costs_carol = [0] * num_techs
previous_costs_gui = [0] * num_techs

def add_current_to_previous_costs(alocation : list[int],this_matrix_costs : list[list[int]], previous_costs : list[int] ):
    for proj, tech in enumerate(alocation):
        previous_costs[tech] += this_matrix_costs[proj][tech]
    return previous_costs

def invert_mtx(mtx : list[list[int]]):
    inv_mtx = []
    for line in mtx:
        new_line = []
        for cell in line:
            new_value = 1 / cell
            new_line.append(new_value)
        inv_mtx.append(new_line)
    return inv_mtx

for month in months_init_projs:
    projs_rows_begin_this_month = projetos_organizados_por_meses[month]
    this_matrix = []
    for proj in projs_rows_begin_this_month:
        this_matrix.append(matrix_aptitude[proj])
    num_projs = len(this_matrix)
    this_matrix_pd = pd.DataFrame(this_matrix)
    alocation = carolina_heuristicaURC(df_aptitude_between_task_machine=this_matrix_pd, num_tasks=num_projs, machine_quantity=num_techs, previous_costs=previous_costs_carol)
    previous_costs_carol = add_current_to_previous_costs(alocation=alocation, this_matrix_costs=this_matrix, previous_costs=previous_costs_carol)
    # this_matrix_pd = pd.DataFrame(invert_mtx(this_matrix))
    alocation = gui_heuristica(df_task_time_per_machine=this_matrix_pd, load_per_machine=previous_costs_gui, num_tasks=num_projs, machine_quantity=num_techs)
    # To join list of lists
    alocation_clean = list(itertools.chain.from_iterable(alocation))
    previous_costs_gui = add_current_to_previous_costs(alocation=alocation_clean, this_matrix_costs=this_matrix, previous_costs=previous_costs_gui)
    print("Alocation month: ", month)

def print_stuff(total_costs : list[int]):
    print(total_costs)
    print("STD: ",statistics.pstdev(total_costs) )
    print("Amp: ",max(total_costs) - min(total_costs) )

print("Carol")
print_stuff(total_costs=previous_costs_carol)
print("Gui")
print_stuff(total_costs=previous_costs_gui)
