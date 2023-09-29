from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook

def write_attribution(excel_workbook,  pairs_with_data, sheetname, iterations : int):
    output_sheet = excel_workbook.create_sheet(sheetname)
    
    output_sheet['A1'] = 'Stand. dev.'
    output_sheet['D1'] = 'Amp.'
    output_sheet['G1'] = 'Diferenças.'
    output_sheet['A2'] = 'Gui'
    output_sheet['B2'] = 'Carolina'
    output_sheet['C2'] = 'Renan'
    output_sheet['D2'] = 'Gui'
    output_sheet['E2'] = 'Carolina'
    output_sheet['F2'] = 'Renan'
    output_sheet['G2'] = 'Gui'
    output_sheet['H2'] = 'Carolina'
    output_sheet['I2'] = 'Renan'
    (results_std, results_total_cost, differences) = pairs_with_data
    for iteration in range(iterations):
        #Write Gui
        output_sheet.cell(row=iteration+3, column=1).value = results_std['Gui'][iteration]
        output_sheet.cell(row=iteration+3, column=2).value = results_std['Carolina'][iteration]
        output_sheet.cell(row=iteration+3, column=3).value = results_std['Renan'][iteration]

        output_sheet.cell(row=iteration+3, column=4).value = results_total_cost['Gui'][iteration]
        output_sheet.cell(row=iteration+3, column=5).value = results_total_cost['Carolina'][iteration]
        output_sheet.cell(row=iteration+3, column=6).value = results_total_cost['Renan'][iteration]

        output_sheet.cell(row=iteration+3, column=7).value = differences['Gui'][iteration]
        output_sheet.cell(row=iteration+3, column=8).value = differences['Carolina'][iteration]
        output_sheet.cell(row=iteration+3, column=9).value = differences['Renan'][iteration]
# The output to be written is a list of ints, saying what technician do what job.
def write_output(pair_small, pair_medium, pair_large, iteration : int):
 
    print("Starting writing output")
    output_excel = Workbook()
    
    #write_attribution(excel_workbook=output_excel, attribution=output)

    write_attribution(excel_workbook=output_excel, pairs_with_data=pair_small, sheetname="small", iterations=iteration) 
    write_attribution(excel_workbook=output_excel, pairs_with_data=pair_medium, sheetname="medium", iterations=iteration) 
    write_attribution(excel_workbook=output_excel, pairs_with_data=pair_large, sheetname="large", iterations=iteration) 
    output_excel.save("./AQUI2.xlsx")
